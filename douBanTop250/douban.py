import random
import re

import requests
from bs4 import BeautifulSoup
import openpyxl
from fake_useragent import UserAgent

user_agent = UserAgent().random

wb = openpyxl.Workbook()
sheet = wb.active

sheet['A1'] = 'Id'
sheet['B1'] = 'Image'
sheet['C1'] = 'Link'
sheet['D1'] = 'Title'
sheet['E1'] = 'Description'
sheet['F1'] = 'Rating'
sheet['G1'] = 'Quote'

row = 2


# 获取随机User_Agent伪装
def get_fake_User_Agent():
    # 随机获取User_Agent
    ua = UserAgent()
    user_anget = ua.random
    return user_anget


# 获取IP伪装
def get_fake_IP():
    ip_page = requests.get(  # 获取200条IP
        'http://www.89ip.cn/tqdl.html?num=60&address=&kill_address=&port=&kill_port=&isp=')
    proxies_list = re.findall(
        r'(25[0-5]|2[0-4]\d|[0-1]\d{2}|[1-9]?\d)\.(25[0-5]|2[0-4]\d|[0-1]\d{2}|[1-9]?\d)\.(25[0-5]|2[0-4]\d|[0-1]\d{2}|[1-9]?\d)\.(25[0-5]|2[0-4]\d|[0-1]\d{2}|[1-9]?\d)(:-?[1-9]\d*)',
        ip_page.text)

    # 转换proxies_list的元素为list,最初为'tuple'元组格式
    proxies_list = list(map(list, proxies_list))

    # 格式化ip  ('112', '111', '217', '188', ':9999')  --->  112.111.217.188:9999
    for u in range(0, len(proxies_list)):
        # 通过小数点来连接为字符
        proxies_list[u] = '.'.join(proxies_list[u])
        # 用rindex()查找最后一个小数点的位置，
        index = proxies_list[u].rindex('.')
        # 将元素转换为list格式
        proxies_list[u] = list(proxies_list[u])
        # 修改位置为index的字符为空白（去除最后一个小数点）
        proxies_list[u][index] = ''
        # 重新通过空白符连接为字符
        proxies_list[u] = ''.join(proxies_list[u])

    # proxies = {'协议':'协议://IP:端口号'}
    # 'https':'https://59.172.27.6:38380'

    return "'" + random.choice(proxies_list) + "'"


def request_douban(url):
    headers = {
        # 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36 Edg/112.0.1722.48',
        'User-Agent': get_fake_User_Agent()
    }
    proxies = {'http': get_fake_IP()}
    try:
        response = requests.get(url=url, proxies=proxies, headers=headers)
        if response.status_code == 200:
            return response.text
    except requests.RequestException:
        return None


def turn_page(page: int):
    url = 'https://movie.douban.com/top250?start={}&filter='.format(page * 25)
    html = request_douban(url)
    soup = BeautifulSoup(html, 'lxml')
    return soup


def get_movie_info(soup: BeautifulSoup):
    # bgCard = soup.find('ol', attrs={'class': 'grid_view'})
    # movies = bgCard.findAll('li')

    movies = soup.find('ol', attrs={'class': 'grid_view'})
    if movies is not None:
        movies = movies.findAll('li')
        # Rest of your code to process the movies list
    else:
        print("No movies found.")

    global row
    moviesList = []
    for movie in movies:
        movieList = []
        try:
            img = movie.find('img').get('src')
        except:
            img = None

        try:
            link = movie.find('div', attrs={'class': 'hd'}).find('a').get('href')
        except:
            link = None
        try:
            title = movie.find('div', attrs={'class': 'hd'}).find('span').string
        except:
            title = None
        try:
            desc = movie.find('div', attrs={'class': 'bd'}).find('p').text
        except:
            desc = None
        try:
            rating = movie.find('div', attrs={'class': 'star'}).find('span', attrs={'class': 'rating_num'}).string
        except:
            rating = None

        try:
            quote = movie.find('p', attrs={'class': 'quote'}).find('span').string
        except:
            quote = None

        movieList.append(img)
        movieList.append(link)
        movieList.append(title)
        movieList.append(desc)
        movieList.append(rating)
        movieList.append(quote)
        moviesList.append(movieList)

        sheet.cell(row=row, column=1, value=row - 1)
        sheet.cell(row=row, column=2, value=img)
        sheet.cell(row=row, column=3, value=link)
        sheet.cell(row=row, column=4, value=title)
        sheet.cell(row=row, column=5, value=desc)
        sheet.cell(row=row, column=6, value=rating)
        sheet.cell(row=row, column=7, value=quote)
        row += 1
        # print(img, title, link, desc, rating, quote)
    # print(moviesList)
    # print(len(moviesList))
    return moviesList


def getAllMovieList(pages: int):
    allMoviesList = []
    for i in range(0, pages):
        soup = turn_page(i)
        movie_infoList = get_movie_info(soup)
        allMoviesList.extend(movie_infoList)
    return allMoviesList


if __name__ == '__main__':
    # for i in range(0, 10):
    #     soup = turn_page(i)
    #     get_movie_info(soup)
    getAllMovieList(10)
    try:
        wb.save('movie.xlsx')
        print('Save success')
    except:
        print('Save failed')
